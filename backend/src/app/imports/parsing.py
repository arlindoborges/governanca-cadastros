from __future__ import annotations

import io
from collections.abc import Callable, Iterator
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

MAX_SAMPLE_ROWS = 5
ZIP_SIGNATURE = b"PK\x03\x04"

ParseProgressCallback = Callable[[int, int], None]


def looks_like_zip(prefix: bytes) -> bool:
    return prefix.startswith(ZIP_SIGNATURE)


def parse_headers_and_rows(
    content: bytes,
    *,
    on_progress: ParseProgressCallback | None = None,
) -> tuple[list[str], list[dict[str, str]]]:
    if not content:
        raise ValueError("XLSX_EMPTY")
    if not looks_like_zip(content[:4]):
        raise ValueError("XLSX_INVALID")

    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except InvalidFileException as exc:
        raise ValueError("XLSX_INVALID") from exc

    try:
        worksheet = workbook.active
        if worksheet is None:
            raise ValueError("XLSX_EMPTY")

        row_iter = worksheet.iter_rows(values_only=True)
        try:
            header_row = next(row_iter)
        except StopIteration as exc:
            raise ValueError("XLSX_EMPTY") from exc

        headers = [_normalize_header(value) for value in header_row]
        while headers and not headers[-1]:
            headers.pop()
        if not headers or any(not header for header in headers):
            raise ValueError("XLSX_HEADER")
        if len(headers) != len(set(headers)):
            raise ValueError("XLSX_DUPLICATE_HEADER")

        estimated_total = max((worksheet.max_row or 1) - 1, 1)
        rows: list[dict[str, str]] = []
        for raw_row in row_iter:
            if raw_row is None:
                continue
            if not any(_cell_has_value(value) for value in raw_row):
                continue
            row = {
                headers[index]: _normalize_cell(raw_row[index])
                for index in range(len(headers))
                if index < len(raw_row)
            }
            rows.append(row)
            if on_progress and (len(rows) == 1 or len(rows) % 500 == 0):
                on_progress(len(rows), estimated_total)

        if not rows:
            raise ValueError("XLSX_NO_DATA")
        if on_progress:
            on_progress(len(rows), max(len(rows), 1))
        return headers, rows
    finally:
        workbook.close()


def preview_rows(rows: list[dict[str, str]], limit: int = MAX_SAMPLE_ROWS) -> list[dict[str, str]]:
    return rows[:limit]


def iter_mapped_values(
    rows: list[dict[str, str]], mapping: dict[str, str]
) -> Iterator[tuple[int, dict[str, str], str | None, str | None, str | None]]:
    code_col = mapping["source_code"]
    description_col = mapping["original_description"]
    unit_col = mapping["original_unit"]
    for index, row in enumerate(rows, start=2):
        yield (
            index,
            row,
            _optional_text(row.get(code_col), 255),
            _optional_text(row.get(description_col), None),
            _optional_text(row.get(unit_col), 100),
        )


def row_issues(
    source_code: str | None,
    original_description: str | None,
    original_unit: str | None,
) -> list[str]:
    issues: list[str] = []
    if source_code is None:
        issues.append("MISSING_SOURCE_CODE")
    if original_description is None:
        issues.append("MISSING_DESCRIPTION")
    if original_unit is None:
        issues.append("MISSING_UNIT")
    return issues


def mapping_from_headers(headers: list[str], payload: dict[str, str]) -> dict[str, str]:
    required = ("source_code", "original_description", "original_unit")
    mapping: dict[str, str] = {}
    header_set = set(headers)
    for field in required:
        column = (payload.get(field) or "").strip()
        if not column:
            raise KeyError(field)
        if column not in header_set:
            raise LookupError(field)
        mapping[field] = column
    if len(set(mapping.values())) != 3:
        raise ValueError("XLSX_MAPPING_OVERLAP")
    return mapping


def _normalize_header(name: Any) -> str:
    return str(name or "").strip()


def _normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _cell_has_value(value: Any) -> bool:
    if value is None:
        return False
    return str(value).strip() != ""


def _optional_text(value: str | None, max_length: int | None) -> str | None:
    text = (value or "").strip()
    if not text:
        return None
    if max_length is not None and len(text) > max_length:
        return None
    return text
