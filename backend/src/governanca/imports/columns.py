from __future__ import annotations

import re
import unicodedata
from io import BytesIO

from openpyxl import load_workbook

FIELD_ALIASES: dict[str, list[str]] = {
    "source_code": [
        "codigo",
        "codigo_prd",
        "codigo_produto",
        "codigo_novo",
        "codigo_item",
        "code",
        "sku",
        "cod",
    ],
    "original_description": [
        "descricao",
        "descricao_original",
        "description",
        "nome",
        "produto",
        "descricao_produto",
        "descr",
        "item",
        "material",
    ],
    "original_unit": [
        "unidade",
        "unit",
        "und",
        "um",
        "uom",
        "unidade_medida",
    ],
}


def normalize_header(value: str) -> str:
    text = unicodedata.normalize("NFKD", value)
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = text.lower().strip()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def read_spreadsheet_headers(file_bytes: bytes) -> list[str]:
    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    if not rows:
        return []
    return [str(cell).strip() if cell is not None else "" for cell in rows[0]]


def build_header_maps(headers: list[str]) -> tuple[dict[str, int], dict[str, int]]:
    exact: dict[str, int] = {}
    normalized: dict[str, int] = {}
    for index, header in enumerate(headers):
        if not header:
            continue
        exact[header] = index
        normalized[normalize_header(header)] = index
    return exact, normalized


def resolve_column_index(
    field: str,
    mapping: dict[str, str],
    exact: dict[str, int],
    normalized: dict[str, int],
) -> int | None:
    mapped = mapping.get(field)
    if mapped:
        if mapped in exact:
            return exact[mapped]
        mapped_norm = normalize_header(mapped)
        if mapped_norm in normalized:
            return normalized[mapped_norm]

    for alias in FIELD_ALIASES.get(field, []):
        if alias in normalized:
            return normalized[alias]
    return None


def suggest_mapping(headers: list[str]) -> dict[str, str]:
    _, normalized = build_header_maps(headers)
    reverse = {index: header for header, index in build_header_maps(headers)[0].items()}
    suggested: dict[str, str] = {}

    for field, aliases in FIELD_ALIASES.items():
        for alias in aliases:
            if alias in normalized:
                suggested[field] = reverse[normalized[alias]]
                break
    return suggested


def cell_value(row: tuple, column_index: int | None) -> str | None:
    if column_index is None or column_index >= len(row):
        return None
    value = row[column_index]
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def iter_importable_records(file_bytes: bytes, mapping: dict[str, str]):
    """Ignora linhas vazias do Excel (faixa formatada além dos dados reais)."""
    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    headers = [str(cell).strip() if cell is not None else "" for cell in header_row]
    exact_headers, normalized_headers = build_header_maps(headers)

    description_index = resolve_column_index(
        "original_description",
        mapping,
        exact_headers,
        normalized_headers,
    )
    if description_index is None:
        return

    code_index = resolve_column_index("source_code", mapping, exact_headers, normalized_headers)
    unit_index = resolve_column_index("original_unit", mapping, exact_headers, normalized_headers)

    for excel_row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        description = cell_value(row, description_index)
        if not description:
            continue
        yield excel_row_num, {
            "source_code": cell_value(row, code_index),
            "original_description": description,
            "original_unit": cell_value(row, unit_index),
        }


def count_importable_records(file_bytes: bytes, mapping: dict[str, str]) -> int:
    return sum(1 for _ in iter_importable_records(file_bytes, mapping))
